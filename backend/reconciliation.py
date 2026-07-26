"""Reconciliation ingestion.

A GST notice is mostly about numbers. The real work on an ASMT-10 is
reconciling GSTR-2A against GSTR-3B supplier by supplier, and only then
arguing the law over what is left. A panel that never sees the figures argues
the difference as one undifferentiated number — which is exactly what a reply
must not do, because a difference is several problems wearing one figure and
each carries a different argument.

THE ARCHITECTURAL DECISION HERE: none of this data reaches a model.

A reconciliation is thousands of rows of client and third-party invoice
detail. Sending it would cost hundreds of thousands of tokens, expose supplier
data that is not even the client's to disclose, and buy nothing — bucketing is
arithmetic, and arithmetic belongs in Python. Rows are parsed, classified and
aggregated locally; what reaches the panel is a summary of a few hundred
tokens: bucket, amount, count, share.

The one exception is deliberate and narrow. Column headers vary between firms
("GSTIN", "Supplier GSTIN", "GSTIN of Supplier", "Party GSTIN"), so when fuzzy
matching cannot identify the columns, the HEADER ROW ALONE may be sent for
mapping. Headers carry no client data. No row ever follows them.
"""

import csv
import io
import os
import re
from typing import Any, Dict, List, Optional, Tuple

MAX_UPLOAD_BYTES = int(os.getenv("MAX_RECON_BYTES", str(15 * 1024 * 1024)))
SUPPORTED = (".xlsx", ".xlsm", ".csv")

# Guard against a workbook with a runaway row count.
MAX_ROWS = 20000

# Rows below this are noise in a reconciliation and distort the buckets.
MIN_MATERIAL_AMOUNT = 1.0


# ---------------------------------------------------------------------------
# Column detection
# ---------------------------------------------------------------------------

# Ordered: the first alias that matches a header wins, so more specific
# aliases must precede generic ones.
COLUMN_ALIASES: Dict[str, List[str]] = {
    "supplier_gstin": [
        "supplier gstin", "gstin of supplier", "vendor gstin", "party gstin",
        "supplier gst no", "gstin/uin", "gstin uin", "gstin",
    ],
    "supplier_name": [
        "supplier name", "vendor name", "party name", "trade name",
        "legal name", "supplier", "vendor", "party",
    ],
    "invoice_no": [
        "invoice number", "invoice no", "inv no", "bill no", "document number",
        "doc no", "invoice",
    ],
    "invoice_date": [
        "invoice date", "inv date", "bill date", "document date", "date",
    ],
    "amount_2a": [
        "2a amount", "gstr-2a", "gstr 2a", "2b amount", "gstr-2b", "gstr 2b",
        "as per 2a", "as per 2b", "as per gstr-2a", "as per gstr-2b",
        "portal amount", "2a tax", "2b tax",
    ],
    "amount_books": [
        "books amount", "as per books", "as per 3b", "3b amount", "gstr-3b",
        "gstr 3b", "purchase register", "as per pr", "books", "itc availed",
        "credit availed",
    ],
    "difference": [
        "difference", "diff", "variance", "short", "excess", "gap",
        "mismatch amount", "net difference",
    ],
    "tax_amount": [
        "total tax", "tax amount", "igst", "cgst", "sgst", "tax", "amount",
    ],
    "status": [
        "status", "remarks", "remark", "reason", "category", "observation",
        "nature of difference", "reconciliation status", "comments", "narration",
    ],
}


def _normalise_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", str(value or "").strip().lower()).strip()


def detect_columns(headers: List[Any]) -> Dict[str, int]:
    """
    Map our field names onto column indices, locally.

    Exact alias match first across all fields, then containment. Doing exact
    matching for everything before any containment stops a loose alias in one
    field stealing a column another field would have matched precisely.
    """
    normalised = [_normalise_header(h) for h in headers]
    mapping: Dict[str, int] = {}
    taken: set = set()

    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for index, header in enumerate(normalised):
                if index in taken or not header:
                    continue
                if header == alias:
                    mapping[field] = index
                    taken.add(index)
                    break
            if field in mapping:
                break

    for field, aliases in COLUMN_ALIASES.items():
        if field in mapping:
            continue
        for alias in aliases:
            for index, header in enumerate(normalised):
                if index in taken or not header:
                    continue
                if alias in header or header in alias:
                    mapping[field] = index
                    taken.add(index)
                    break
            if field in mapping:
                break

    return mapping


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------


def _to_amount(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    # Accounting parentheses are the usual way a credit is shown in a
    # reconciliation exported from Tally.
    negative = text.startswith("(") and text.endswith(")")
    # Indian digit grouping, and a currency prefix that would otherwise leave
    # a stray decimal point ahead of the number ("Rs. 1,000.50").
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    amount = float(match.group())
    return -abs(amount) if negative else amount


def _find_header_row(rows: List[List[Any]]) -> int:
    """
    Reconciliation sheets routinely carry a title and a blank line above the
    real header. Take the first row within the first fifteen that maps at
    least two known fields.
    """
    best_index, best_score = 0, 0
    for index, row in enumerate(rows[:15]):
        score = len(detect_columns(row))
        if score > best_score:
            best_index, best_score = index, score
        if score >= 3:
            return index
    return best_index if best_score >= 2 else 0


def parse_workbook(filename: str, content: bytes) -> Tuple[List[Any], List[List[Any]], List[str]]:
    """
    Read an uploaded reconciliation into headers and rows. Entirely local.

    Returns (headers, rows, warnings).
    """
    warnings: List[str] = []
    name = (filename or "").lower()

    if len(content) > MAX_UPLOAD_BYTES:
        raise ValueError(
            f"File is larger than {MAX_UPLOAD_BYTES // (1024 * 1024)} MB."
        )
    if not any(name.endswith(ext) for ext in SUPPORTED):
        raise ValueError(
            f"Unsupported file type. Upload one of: {', '.join(SUPPORTED)}."
        )

    raw: List[List[Any]] = []
    if name.endswith(".csv"):
        text = None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                text = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("Could not decode the CSV file.")
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        raw = [row for row in csv.reader(io.StringIO(text), dialect)]
    else:
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
            sheet = workbook.active
            for index, row in enumerate(sheet.iter_rows(values_only=True)):
                if index > MAX_ROWS:
                    warnings.append(
                        f"Only the first {MAX_ROWS:,} rows were read."
                    )
                    break
                raw.append(list(row))
            workbook.close()
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(f"Could not read the workbook: {e}")

    raw = [row for row in raw if any(
        cell not in (None, "") for cell in (row or [])
    )]
    if not raw:
        raise ValueError("The file contains no data.")

    header_index = _find_header_row(raw)
    headers = raw[header_index]
    rows = raw[header_index + 1:]

    if not rows:
        warnings.append("A header row was found but there are no data rows.")

    return headers, rows, warnings


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------


def classify_row(
    row: List[Any],
    mapping: Dict[str, int],
    pack,
) -> str:
    """
    Assign a line to a legal bucket.

    Preference order is deliberate. What the preparer wrote in the status
    column is the best evidence of what the difference actually is — they did
    the reconciliation. Only where that is silent do we fall back to the shape
    of the row, and where neither speaks the line is left UNRECONCILED rather
    than assigned a flattering category.
    """
    def cell(field: str) -> str:
        index = mapping.get(field)
        if index is None or index >= len(row):
            return ""
        return str(row[index] or "").strip()

    status = cell("status").lower()
    if status:
        for bucket in pack.RECONCILIATION_BUCKETS:
            if any(keyword in status for keyword in bucket.keywords):
                return bucket.key

    # The supplier's own name sometimes carries the answer ("RCM - freight").
    haystack = f"{cell('supplier_name')} {cell('invoice_no')}".lower()
    for key in ("rcm", "import_igst", "isd"):
        bucket = pack.RECONCILIATION_BUCKETS_BY_KEY[key]
        if any(keyword in haystack for keyword in bucket.keywords):
            return bucket.key

    # A line present in the books but absent from the portal, with no
    # explanation, is unexplained. It is NOT assumed to be timing.
    return pack.UNRECONCILED.key


def row_amount(row: List[Any], mapping: Dict[str, int]) -> Optional[float]:
    """
    The amount in issue for a line.

    An explicit difference column is authoritative. Otherwise the difference
    between books and portal. Otherwise whatever single amount is present.
    """
    def value(field: str) -> Optional[float]:
        index = mapping.get(field)
        if index is None or index >= len(row):
            return None
        return _to_amount(row[index])

    difference = value("difference")
    if difference is not None:
        return abs(difference)

    books, portal = value("amount_books"), value("amount_2a")
    if books is not None and portal is not None:
        return abs(books - portal)
    for field in ("amount_books", "amount_2a", "tax_amount"):
        amount = value(field)
        if amount is not None:
            return abs(amount)
    return None


def summarise(
    headers: List[Any],
    rows: List[List[Any]],
    pack,
    mapping: Optional[Dict[str, int]] = None,
    mask_suppliers: bool = False,
) -> Dict[str, Any]:
    """
    Bucket the reconciliation and aggregate.

    The return value is what the panel sees. It contains totals, counts and
    shares — never a row.
    """
    mapping = mapping if mapping is not None else detect_columns(headers)
    warnings: List[str] = []

    if not mapping:
        raise ValueError(
            "None of the columns could be identified. Expected at least a "
            "supplier GSTIN or an amount column."
        )
    if not any(k in mapping for k in
               ("difference", "amount_books", "amount_2a", "tax_amount")):
        raise ValueError(
            "No amount column could be identified. Expected a difference, "
            "books, portal or tax amount column."
        )
    if "status" not in mapping:
        warnings.append(
            "No status or remarks column was found, so lines could not be "
            "classified from the preparer's own reconciliation and are shown "
            "as not yet reconciled. Add a remarks column describing each "
            "difference for a materially better analysis."
        )

    buckets: Dict[str, Dict[str, Any]] = {}
    suppliers: Dict[str, float] = {}
    total = 0.0
    counted = 0
    skipped = 0

    for row in rows:
        amount = row_amount(row, mapping)
        if amount is None or amount < MIN_MATERIAL_AMOUNT:
            skipped += 1
            continue

        key = classify_row(row, mapping, pack)
        entry = buckets.setdefault(key, {"key": key, "amount": 0.0, "count": 0})
        entry["amount"] += amount
        entry["count"] += 1
        total += amount
        counted += 1

        gstin_index = mapping.get("supplier_gstin")
        if gstin_index is not None and gstin_index < len(row):
            gstin = str(row[gstin_index] or "").strip()
            if gstin:
                suppliers[gstin] = suppliers.get(gstin, 0.0) + amount

    if not counted:
        raise ValueError(
            "No usable amounts were found. Check that the amount column "
            "contains numbers."
        )

    ordered = sorted(buckets.values(), key=lambda b: b["amount"], reverse=True)
    for entry in ordered:
        entry["share"] = entry["amount"] / total if total else 0.0
        bucket = (pack.RECONCILIATION_BUCKETS_BY_KEY.get(entry["key"])
                  or pack.UNRECONCILED)
        entry["label"] = bucket.label
        entry["strength"] = bucket.strength
        entry["action"] = bucket.action

    unreconciled = next(
        (b for b in ordered if b["key"] == pack.UNRECONCILED.key), None
    )
    if unreconciled and unreconciled["share"] > 0.25:
        warnings.append(
            f"{unreconciled['share']:.0%} of the difference "
            f"(Rs. {unreconciled['amount']:,.0f}) is unexplained. That portion "
            "has no argument behind it — trace it before filing."
        )

    top = sorted(suppliers.items(), key=lambda item: item[1], reverse=True)[:5]
    exposures = [
        {
            # Supplier GSTINs are third-party data and not the client's to
            # disclose. On the anonymising tier they never leave the machine.
            "supplier": ("[supplier withheld]" if mask_suppliers
                         else gstin),
            "amount": round(amount, 2),
        }
        for gstin, amount in top
    ]

    return {
        "total": round(total, 2),
        "row_count": counted,
        "skipped_rows": skipped,
        "supplier_count": len(suppliers),
        "buckets": ordered,
        "top_exposures": exposures,
        "columns_detected": sorted(mapping.keys()),
        "warnings": warnings,
    }


def read_reconciliation(
    filename: str,
    content: bytes,
    pack,
    mask_suppliers: bool = False,
) -> Dict[str, Any]:
    """Parse, classify and aggregate an uploaded reconciliation."""
    headers, rows, warnings = parse_workbook(filename, content)
    summary = summarise(headers, rows, pack, mask_suppliers=mask_suppliers)
    summary["warnings"] = warnings + summary["warnings"]
    summary["headers"] = [str(h) for h in headers if h not in (None, "")][:40]
    return summary
